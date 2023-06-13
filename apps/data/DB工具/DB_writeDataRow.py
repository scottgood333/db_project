import openpyxl
import psycopg2
from DB_config import POSTGRES

def read_excel_and_write_to_database(excel_file):
    
    # 讀取 Excel 檔案
    workBook = openpyxl.load_workbook(excel_file)
    sheets = workBook.sheetnames
    
    # 遍歷所有工作表 ( 資料表 )
    for sheet_name in sheets:
        # 取得工作表 ( 資料表 )
        sheet = workBook[sheet_name]
        
        # 取得資料庫名稱、模式名稱、資料表名稱
        db_name = sheet['A2'].value
        schema_name = sheet['B2'].value
        table_name = sheet['C2'].value
        
        # 連接到 PostgreSQL 資料庫
        conn = psycopg2.connect(
            user = POSTGRES['user'],
            password = POSTGRES['password'],
            host = POSTGRES['host'],
            port = POSTGRES['port'],
            dbname = db_name
        )

        # 取得「欄位名稱」
        headers = [cell.value for cell in sheet[4] if cell.value is not None]

        ### 建立插入資料的 SQL 語句
        # 把「欄位名稱」變成 SQL 插入語句的一部分
        columns = ', '.join(headers)
        # 根據「欄位名稱」數量預先給定 SQL 插入語句的占位符
        values_placeholder = ', '.join(['%s'] * len(headers))
        # 組合 SQL 插入語句
        SQL_cmd_insert = f'INSERT INTO "{schema_name}"."{table_name}" ({columns}) VALUES ({values_placeholder})'
        
        # 取得資料列的內容並將資料值取代占位符後插入資料庫
        # min_row：從哪一行開始遍歷 / max_ool：行取值取到哪一欄，這邊要配合占位符(資料欄)數量
        # values_only：只返回資料值，有興趣可以查一下如果 False 可以玩出哪些花
        for row in sheet.iter_rows(min_row=5, max_col=len(headers), values_only=True):
            
            # 去除空白資料列
            if all(cell is None for cell in row):
                continue
            ### 執行插入資料的 SQL 語句
            # 產生並指派光標
            cursor = conn.cursor()
            # 將「值們」取代占位符
            cursor.execute(SQL_cmd_insert, row)
            # 結束光標
            cursor.close()
    
        # 提交更改並關閉連接
        conn.commit()
        conn.close()

if __name__ == '__main__':
    # Excel 檔案路徑
    excel_file = 'data.xlsx'
    
    # 讀取 Excel 檔案並將資料寫入資料庫
    read_excel_and_write_to_database(excel_file)
